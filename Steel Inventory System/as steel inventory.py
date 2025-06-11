import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from datetime import datetime
from openpyxl import Workbook, load_workbook
import customtkinter as ctk
import os

class SteelInventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AS Steel Mills")

        # Set up initial variables
        self.parties = []
        self.load_parties()

        # Create main screen
        self.main_screen()

    def load_parties(self):
        # Load existing parties from a file if it exists
        if os.path.exists("parties.txt"):
            with open("parties.txt", "r") as file:
                self.parties = file.read().splitlines()

    def save_parties(self):
        # Save parties to a file
        with open("parties.txt", "w") as file:
            file.write("\n".join(self.parties))

    def main_screen(self):
        self.clear_screen()
        
        title = ctk.CTkLabel(self.root, text="AS Steel Mills", font=("Helvetica", 24, "bold"))
        title.pack(pady=50)

        enter_data_button = ctk.CTkButton(self.root, text="Enter New Data", command=self.data_entry_screen, width=200)
        enter_data_button.pack(pady=20)

        open_ledger_button = ctk.CTkButton(self.root, text="Open Ledger", command=self.open_ledger_screen, width=200)
        open_ledger_button.pack(pady=20)

    def data_entry_screen(self):
        self.clear_screen()

        title = ctk.CTkLabel(self.root, text="AS Steel Inventory", font=("Helvetica", 20, "bold"))
        title.pack(pady=10)

        # Frame for party selection
        self.party_frame = ctk.CTkFrame(self.root)
        self.party_frame.pack(padx=20, pady=10, fill=tk.X)

        self.party_label = ctk.CTkLabel(self.party_frame, text="Enter a Party:")
        self.party_label.pack(side=tk.LEFT, padx=10)

        self.party_var = tk.StringVar()
        self.party_dropdown = ttk.Combobox(self.party_frame, values=self.parties, textvariable=self.party_var)
        self.party_dropdown.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        self.party_dropdown.bind('<KeyRelease>', self.check_input)

        self.add_party_button = ctk.CTkButton(self.party_frame, text="Add Party", command=self.add_party_popup, fg_color="green")
        self.add_party_button.pack(side=tk.LEFT, padx=10)

        self.remove_party_button = ctk.CTkButton(self.party_frame, text="Remove Party", command=self.remove_party_popup, fg_color="red")
        self.remove_party_button.pack(side=tk.LEFT, padx=10)

        # Frame for action selection
        self.action_frame = ctk.CTkFrame(self.root)
        self.action_frame.pack(padx=20, pady=10, fill=tk.X)

        self.action_label = ctk.CTkLabel(self.action_frame, text="Action:")
        self.action_label.pack(side=tk.LEFT, padx=10)

        self.action_var = tk.StringVar()

        self.in_button = ctk.CTkButton(self.action_frame, text="IN", command=lambda: self.set_action("IN"), width=80)
        self.in_button.pack(side=tk.LEFT, padx=10)

        self.out_button = ctk.CTkButton(self.action_frame, text="OUT", command=lambda: self.set_action("OUT"), width=80)
        self.out_button.pack(side=tk.LEFT, padx=10)

        # Frame for date selection
        self.date_frame = ctk.CTkFrame(self.root)
        self.date_frame.pack(padx=20, pady=10, fill=tk.X)

        self.date_label = ctk.CTkLabel(self.date_frame, text="Date:")
        self.date_label.pack(side=tk.LEFT, padx=10)

        self.current_date_button = ctk.CTkButton(self.date_frame, text="Current Date", command=self.use_current_date, width=120)
        self.current_date_button.pack(side=tk.LEFT, padx=10)

        self.custom_date_button = ctk.CTkButton(self.date_frame, text="Custom Date", command=self.use_custom_date, width=120)
        self.custom_date_button.pack(side=tk.LEFT, padx=10)

        self.custom_date_var = tk.StringVar()
        self.custom_date_entry = ctk.CTkEntry(self.date_frame, textvariable=self.custom_date_var)

        # Frame for weight input
        self.weight_frame = ctk.CTkFrame(self.root)
        self.weight_frame.pack(padx=20, pady=10, fill=tk.X)

        self.weight_label = ctk.CTkLabel(self.weight_frame, text="Enter Weight:")
        self.weight_label.pack(side=tk.LEFT, padx=10)

        self.weight_var = tk.IntVar()
        self.weight_entry = ctk.CTkEntry(self.weight_frame, textvariable=self.weight_var)
        self.weight_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

        # Frame for car number input
        self.car_frame = ctk.CTkFrame(self.root)
        self.car_frame.pack(padx=20, pady=10, fill=tk.X)

        self.car_label = ctk.CTkLabel(self.car_frame, text="Enter Car No.:")
        self.car_label.pack(side=tk.LEFT, padx=10)

        self.car_var = tk.StringVar()
        self.car_entry = ctk.CTkEntry(self.car_frame, textvariable=self.car_var)
        self.car_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

        # Frame for towards party input
        self.towards_frame = ctk.CTkFrame(self.root)
        self.towards_frame.pack(padx=20, pady=10, fill=tk.X)

        self.towards_label = ctk.CTkLabel(self.towards_frame, text="Towards Party:")
        self.towards_label.pack(side=tk.LEFT, padx=10)

        self.towards_button = ctk.CTkCheckBox(self.towards_frame, text="Enter Towards Party", command=self.toggle_towards_entry)
        self.towards_button.pack(side=tk.LEFT, padx=10)

        self.towards_var = tk.StringVar()
        self.towards_entry = ctk.CTkEntry(self.towards_frame, textvariable=self.towards_var)

        # Frame for description input
        self.description_frame = ctk.CTkFrame(self.root)
        self.description_frame.pack(padx=20, pady=10, fill=tk.X)

        self.description_label = ctk.CTkLabel(self.description_frame, text="Description:")
        self.description_label.pack(side=tk.LEFT, padx=10)

        self.description_button = ctk.CTkCheckBox(self.description_frame, text="Enter Description", command=self.toggle_description_entry)
        self.description_button.pack(side=tk.LEFT, padx=10)

        self.description_var = tk.StringVar()
        self.description_entry = ctk.CTkEntry(self.description_frame, textvariable=self.description_var)

        # Submit button
        self.submit_button = ctk.CTkButton(self.root, text="Submit", command=self.submit)
        self.submit_button.pack(pady=20)

        # Back button
        self.back_button = ctk.CTkButton(self.root, text="Back", command=self.main_screen)
        self.back_button.pack(pady=10)

    def open_ledger_screen(self):
        self.clear_screen()
        
        title = ctk.CTkLabel(self.root, text="AS Steel Inventory Ledger", font=("Helvetica", 20, "bold"))
        title.pack(pady=10)

        if not self.parties:
            messagebox.showinfo("No Parties", "No parties available. Please add parties through the data entry screen.")
            self.main_screen()
            return
        
        # Display ledger options
        ledger_frame = ctk.CTkFrame(self.root)
        ledger_frame.pack(pady=10)

        in_button = ctk.CTkButton(ledger_frame, text="IN", command=lambda: self.view_ledger("IN"), width=100)
        in_button.pack(side=tk.LEFT, padx=10)

        out_button = ctk.CTkButton(ledger_frame, text="OUT", command=lambda: self.view_ledger("OUT"), width=100)
        out_button.pack(side=tk.LEFT, padx=10)

        for party in self.parties:
            button = ctk.CTkButton(ledger_frame, text=party, command=lambda p=party: self.view_ledger(p), width=100)
            button.pack(side=tk.LEFT, padx=10)

        # Back button
        self.back_button = ctk.CTkButton(self.root, text="Back", command=self.main_screen)
        self.back_button.pack(pady=10)

    def clear_screen(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def set_action(self, action):
        self.action_var.set(action)
        if action == "IN":
            self.in_button.configure(fg_color="green")
            self.out_button.configure(fg_color="default")
        else:
            self.in_button.configure(fg_color="default")
            self.out_button.configure(fg_color="red")

    def use_current_date(self):
        self.custom_date_entry.pack_forget()
        self.current_date_button.configure(fg_color="blue")
        self.custom_date_button.configure(fg_color="default")

    def use_custom_date(self):
        self.custom_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        self.current_date_button.configure(fg_color="default")
        self.custom_date_button.configure(fg_color="blue")

    def toggle_towards_entry(self):
        if self.towards_button.get():
            self.towards_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        else:
            self.towards_entry.pack_forget()

    def toggle_description_entry(self):
        if self.description_button.get():
            self.description_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        else:
            self.description_entry.pack_forget()

    def check_input(self, event):
        value = event.widget.get()
        if value == "":
            self.party_dropdown['values'] = self.parties
        else:
            data = []
            for item in self.parties:
                if value.lower() in item.lower():
                    data.append(item)
            self.party_dropdown['values'] = data

    def add_party_popup(self):
        new_party = simpledialog.askstring("Add Party", "Enter new party name:")
        if new_party and new_party not in self.parties:
            self.parties.append(new_party)
            self.party_dropdown['values'] = self.parties
            self.party_var.set("")
            self.save_parties()

    def remove_party_popup(self):
        party_to_remove = simpledialog.askstring("Remove Party", "Enter party name to remove:")
        if party_to_remove in self.parties:
            self.parties.remove(party_to_remove)
            self.party_dropdown['values'] = self.parties
            self.party_var.set("")
            self.save_parties()

    def submit(self):
        party = self.party_var.get().strip()
        action = self.action_var.get().strip()
        date = datetime.now().strftime("%Y-%m-%d") if not self.custom_date_var.get().strip() else self.custom_date_var.get().strip()

        weight_input = str(self.weight_var.get()).strip()
        if weight_input.isdigit():
            weight = int(weight_input)
        else:
            weight = 0

        car_no = self.car_var.get().strip()
        towards_party = self.towards_var.get().strip() if self.towards_button.get() else "N/A"
        description = self.description_var.get().strip() if self.description_button.get() else "N/A"

        if not party or not action or weight <= 0 or not car_no:
            messagebox.showwarning("Missing Information", "Please fill out all required fields correctly.\nWeight must be a number greater than 0.")
            return

        filename = f"{action}_ledger.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Party", "Date", "Weight", "Car No.", "Towards Party", "Description"])
            wb.save(filename)

        wb = load_workbook(filename)
        ws = wb.active
        ws.append([party, date, weight, car_no, towards_party, description])
        wb.save(filename)

        party_filename = f"{party}_ledger.xlsx"
        if not os.path.exists(party_filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Action", "Date", "Weight", "Car No.", "Towards Party", "Description"])
            wb.save(party_filename)

        wb = load_workbook(party_filename)
        ws = wb.active
        ws.append([action, date, weight, car_no, towards_party, description])
        wb.save(party_filename)

        # Reset all fields
        self.party_var.set("")
        self.action_var.set("")
        self.custom_date_var.set("")
        self.weight_var.set("")
        self.car_var.set("")
        self.towards_var.set("")
        self.description_var.set("")
        self.custom_date_entry.pack_forget()
        self.towards_entry.pack_forget()
        self.description_entry.pack_forget()

        messagebox.showinfo("Success", "Data entered successfully.")



    def view_ledger(self, ledger_type):
        filename = f"{ledger_type}_ledger.xlsx" if ledger_type in ["IN", "OUT"] else f"{ledger_type}_ledger.xlsx"
        if not os.path.exists(filename):
            messagebox.showwarning("File Not Found", f"No data found for {ledger_type}.")
            return

        ledger_window = tk.Toplevel(self.root)
        ledger_window.title(f"{ledger_type} Ledger")

        wb = load_workbook(filename)
        ws = wb.active

        cols = ws.max_column
        rows = ws.max_row

        table = ttk.Treeview(ledger_window, columns=[f"#{i}" for i in range(1, cols+1)], show='headings')
        for i in range(1, cols+1):
            table.heading(f"#{i}", text=ws.cell(row=1, column=i).value)
            table.column(f"#{i}", anchor='center')  # Center-align text

        for row in ws.iter_rows(min_row=2, values_only=True):
            table.insert('', 'end', values=row)

        table.pack(fill=tk.BOTH, expand=True)

if __name__ == "__main__":
    root = ctk.CTk()
    app = SteelInventoryApp(root)
    root.mainloop()
